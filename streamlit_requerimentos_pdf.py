import io
import os
import re
import zipfile
from datetime import datetime

import fitz  # PyMuPDF
import openpyxl
import streamlit as st
from pypdf import PdfReader, PdfWriter
from pypdf.generic import BooleanObject, NameObject, TextStringObject

# ==========================================
# CONFIG
# ==========================================
TEMPLATE_PDF_PATH = "modelo.pdf"

st.set_page_config(page_title="Requerimentos", page_icon="📄", layout="wide")
st.title("📄 Gerador de Requerimentos")

# ==========================================
# CAMPOS REAIS DO PDF
# ==========================================
PDF_FIELDS = {
    "cabecalho": {
        "cartorio": "Text1",
    },
    "nascimento": {
        "check": "1",
        "nome": "Nome_1",
        "termo": "TextField",
        "fls": "TextField_1",
        "livro": "TextField_2",
    },
    "casamento": {
        "check": "TextField_3",
        "nome1": "Text3",
        "nome2": "Nome 2",
        "termo": "TextField_4",
        "fls": "TextField_5",
        "livro": "TextField_6",
    },
    "obito": {
        "check": "TextField_7",
        "nome": "Nome_2",
        "termo": "Termo nº",
        "fls": "Fls",
        "livro": "Livro",
    },
    "especificacoes": {
        "digitada": "TextField_16",
        "fotocopia": "TextField_17",
        "duas": "TextField_18",
        "firma_nao": "TextField_20",
        "haia_nao": "Sim(Se positivo o serviço será cobrado)",
    },
    "rodape": {
        "local": "Local",
        "data": "Data",
    },
}

# ==========================================
# UTIL
# ==========================================
def limpar_texto(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def formatar_valor_excel(valor) -> str:
    if valor is None:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")

    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).replace(".", ",")

    return str(valor).strip()


def sanitizar_nome_arquivo(valor: str) -> str:
    valor = limpar_texto(valor)
    valor = re.sub(r'[\\/:*?"<>|]+', "", valor)
    valor = re.sub(r"\s+", "_", valor)
    return valor[:120] if valor else "SemNome"


def normalizar_tipo(valor: str) -> str:
    v = limpar_texto(valor).lower()
    if "casamento" in v:
        return "casamento"
    if "óbito" in v or "obito" in v:
        return "obito"
    if "nascimento" in v:
        return "nascimento"
    raise ValueError(f"Tipo de certidão inválido: {valor}")


def normalizar_formato(valor: str) -> str:
    v = limpar_texto(valor).lower()
    if "digitada" in v:
        return "digitada"
    if "fotocópia" in v or "fotocopia" in v:
        return "fotocopia"
    if "duas" in v:
        return "duas"
    raise ValueError(f"Formato inválido na coluna H: {valor}")


# ==========================================
# LEITURA EXCEL
# A Nome
# B Cartório
# C Certidão
# D Cônjuge
# E Termo nº
# F Fls
# G Livro
# H Formato
# I Local
# J Data
# ==========================================
def carregar_excel(arquivo_excel) -> list:
    wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
    ws = wb.active

    registros = []

    for i in range(2, ws.max_row + 1):
        nome = ws[f"A{i}"].value
        cartorio = ws[f"B{i}"].value
        tipo = ws[f"C{i}"].value
        conjuge = ws[f"D{i}"].value
        termo = ws[f"E{i}"].value
        fls = ws[f"F{i}"].value
        livro = ws[f"G{i}"].value
        formato = ws[f"H{i}"].value
        local = ws[f"I{i}"].value
        data = ws[f"J{i}"].value

        if not any([nome, cartorio, tipo]):
            continue

        registros.append(
            {
                "linha": i,
                "nome": formatar_valor_excel(nome),
                "cartorio": formatar_valor_excel(cartorio),
                "tipo": formatar_valor_excel(tipo),
                "conjuge": formatar_valor_excel(conjuge),
                "termo": formatar_valor_excel(termo),
                "fls": formatar_valor_excel(fls),
                "livro": formatar_valor_excel(livro),
                "formato": formatar_valor_excel(formato),
                "local": formatar_valor_excel(local) or "Rio de Janeiro",
                "data": formatar_valor_excel(data) or datetime.today().strftime("%d/%m/%Y"),
            }
        )

    return registros


# ==========================================
# APARÊNCIA DOS CAMPOS
# ==========================================
def configurar_aparencia_campos(writer: PdfWriter):
    root = writer._root_object

    if "/AcroForm" in root:
        acroform = root["/AcroForm"]
        acroform.update(
            {
                NameObject("/NeedAppearances"): BooleanObject(True),
                NameObject("/DA"): TextStringObject("/Helv 0 Tf 0 g"),
            }
        )


# ==========================================
# MONTAR CAMPOS
# ==========================================
def montar_campos_pdf(registro: dict) -> dict:
    tipo = normalizar_tipo(registro["tipo"])
    formato = normalizar_formato(registro["formato"])

    campos = {
        PDF_FIELDS["cabecalho"]["cartorio"]: registro["cartorio"],
        PDF_FIELDS["rodape"]["local"]: registro["local"],
        PDF_FIELDS["rodape"]["data"]: registro["data"],
        PDF_FIELDS["especificacoes"]["firma_nao"]: "X",
        PDF_FIELDS["especificacoes"]["haia_nao"]: "X",
    }

    if tipo == "nascimento":
        campos.update({
            PDF_FIELDS["nascimento"]["check"]: "X",
            PDF_FIELDS["nascimento"]["nome"]: registro["nome"],
            PDF_FIELDS["nascimento"]["termo"]: registro["termo"],
            PDF_FIELDS["nascimento"]["fls"]: registro["fls"],
            PDF_FIELDS["nascimento"]["livro"]: registro["livro"],
        })

    elif tipo == "casamento":
        campos.update({
            PDF_FIELDS["casamento"]["check"]: "X",
            PDF_FIELDS["casamento"]["nome1"]: registro["nome"],
            PDF_FIELDS["casamento"]["nome2"]: registro["conjuge"],
            PDF_FIELDS["casamento"]["termo"]: registro["termo"],
            PDF_FIELDS["casamento"]["fls"]: registro["fls"],
            PDF_FIELDS["casamento"]["livro"]: registro["livro"],
        })

    elif tipo == "obito":
        campos.update({
            PDF_FIELDS["obito"]["check"]: "X",
            PDF_FIELDS["obito"]["nome"]: registro["nome"],
            PDF_FIELDS["obito"]["termo"]: registro["termo"],
            PDF_FIELDS["obito"]["fls"]: registro["fls"],
            PDF_FIELDS["obito"]["livro"]: registro["livro"],
        })

    if formato == "digitada":
        campos[PDF_FIELDS["especificacoes"]["digitada"]] = "X"
    elif formato == "fotocopia":
        campos[PDF_FIELDS["especificacoes"]["fotocopia"]] = "X"
    else:
        campos[PDF_FIELDS["especificacoes"]["duas"]] = "X"

    return campos


# ==========================================
# GERAR PDF PREENCHIDO
# ==========================================
def gerar_pdf_preenchido(template_bytes: bytes, registro: dict) -> bytes:
    reader = PdfReader(io.BytesIO(template_bytes))
    writer = PdfWriter()

    for page in reader.pages:
        writer.add_page(page)

    if reader.trailer["/Root"].get("/AcroForm"):
        writer._root_object.update(
            {NameObject("/AcroForm"): reader.trailer["/Root"]["/AcroForm"]}
        )

    configurar_aparencia_campos(writer)

    campos = montar_campos_pdf(registro)
    writer.update_page_form_field_values(
        writer.pages[0],
        campos,
        auto_regenerate=True,
    )

    saida = io.BytesIO()
    writer.write(saida)
    return saida.getvalue()


# ==========================================
# "IMPRIMIR" DE VERDADE
# renderiza cada página e cria um PDF novo
# sem campos editáveis
# ==========================================
def imprimir_pdf_virtual(pdf_bytes: bytes, dpi: int = 200) -> bytes:
    origem = fitz.open(stream=pdf_bytes, filetype="pdf")
    destino = fitz.open()

    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)

    for pagina in origem:
        pix = pagina.get_pixmap(matrix=matrix, alpha=False)
        img_bytes = pix.tobytes("png")

        nova = destino.new_page(width=pagina.rect.width, height=pagina.rect.height)
        nova.insert_image(pagina.rect, stream=img_bytes)

    out = destino.tobytes(garbage=4, deflate=True)
    origem.close()
    destino.close()
    return out


def montar_nome_saida(registro: dict) -> str:
    tipo = normalizar_tipo(registro["tipo"]).capitalize()
    nome = sanitizar_nome_arquivo(registro["nome"])
    return f"Requerimento_{tipo}_{nome}.pdf"


# ==========================================
# UI
# ==========================================
file = st.file_uploader("Excel", type=["xlsx"])

if not os.path.exists(TEMPLATE_PDF_PATH):
    st.error("modelo.pdf não encontrado")

elif file:
    with open(TEMPLATE_PDF_PATH, "rb") as f:
        template = f.read()

    dados = carregar_excel(file)

    zip_buffer = io.BytesIO()
    erros = []

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for reg in dados:
            try:
                pdf_preenchido = gerar_pdf_preenchido(template, reg)
                pdf_final = imprimir_pdf_virtual(pdf_preenchido)
                z.writestr(nome_saida(reg), pdf_final)
            except Exception as e:
                erros.append(f"Linha {reg['linha']}: {e}")

    zip_buffer.seek(0)

    if erros:
        st.error("Algumas linhas falharam:")
        for erro in erros:
            st.write(f"- {erro}")

    st.download_button(
        "📦 Baixar PDFs",
        data=zip_buffer.getvalue(),
        file_name="requerimentos.zip",
        mime="application/zip",
        use_container_width=True,
    )

else:
    st.info("Envie a planilha")
